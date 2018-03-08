# -*- coding: UTF-8 -*-
"""

"""
__author__ = 'Steven howtoesc@gmail.com'
import inspect
import subprocess
import re
from functools import wraps

# import xlwings as xw
from lxml import etree
from openpyxl import load_workbook
from common import *

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger('SR')
logger.setLevel(logging.INFO)

from datetime import datetime


class SceneRunnerRError(RuntimeError):
    def __init__(self, message):

        if isinstance(message, unicode):
            super(SceneRunnerRError, self).__init__(message.encode('utf-8'))
            self.message = message

        elif isinstance(message, str):
            super(SceneRunnerRError, self).__init__(message)
            self.message = message.decode('utf-8')

        # This shouldn't happen...
        else:
            raise TypeError

    def __unicode__(self):

        return self.message


class SceneRunnerRWarning(RuntimeWarning):
    def __init__(self, message):

        if isinstance(message, unicode):
            super(SceneRunnerRWarning, self).__init__(message.encode('utf-8'))
            self.message = message

        elif isinstance(message, str):
            super(SceneRunnerRWarning, self).__init__(message)
            self.message = message.decode('utf-8')

        # This shouldn't happen...
        else:
            raise TypeError

    def __unicode__(self):

        return self.message


class SceneRunnerAssertionError(AssertionError):
    def __init__(self, message):

        if isinstance(message, unicode):
            super(SceneRunnerAssertionError, self).__init__(message.encode('utf-8'))
            self.message = message

        elif isinstance(message, str):
            super(SceneRunnerAssertionError, self).__init__(message)
            self.message = message.decode('utf-8')

        # This shouldn't happen...
        else:
            raise TypeError

    def __unicode__(self):

        return self.message


def runner_assert(identity, ass, iss=True, ):
    # type: (object, object, object) -> object
    scene_app.runner_assert(identity, ass, iss, inspect.currentframe())


def runner_logger(msg, log_level=logging.INFO):
    scene_app.runner_logger(msg, log_level)


def ur(s, r, n):
    return s.encode("utf-8").replace(r.encode("utf-8"), n.encode("utf-8")).decode("utf-8")


def find_node_by_xpath(x, xx, time_out=None):
    return scene_app.find_node_by_xpath(x, xx, time_out)


def get_node_parents(n):
    return scene_app.get_node_parents(n)


def get_node_full_path(n):
    return scene_app.get_node_full_path(n)


def node2gameobj(n):
    return scene_app.get_node_full_path(n)


def press_node(n, press_time):
    return scene_app.press_node(n, press_time)


def find_gameobj_by_xpath(x, xx, time_out=None):
    # type: (object, object, object) -> object
    return scene_app.find_gameobj_by_xpath(x, xx, time_out)


def find_and_press(x, xx='', time_out=None, press_time=100):
    return scene_app.find_and_press(x, xx, time_out, press_time)


def find_and_swipe(x, x1, xx='',time_out=None, steps=20, duration=1000):
    # type: (object, object, object, object, object, object) -> object
    # type: (object, object, object, object, object, object) -> object
    return scene_app.find_and_swipe(x, x1, xx,time_out, steps, duration)


def find_and_press_random(x, xx='',time_out=None, press_time=100):
    return scene_app.find_and_press_random(x, xx,time_out, press_time)


def find_and_swipe_random(x, x1, time_out=None, steps=20, duration=1000):
    return scene_app.find_and_swipe_random(x, x1, time_out, steps, duration)


class SceneRunner:
    def __init__(self):
        self.device = ''

        self.engine = None

        self.SceneModules = {}
        self.Case = {}
        self.CaseTree = {}
        self.Record = {}

        self.dump = None
        self.scene_module = ''
        self.scenes = {}
        self.scene_now = ''
        self.last_scene = ''

        self.timer = time.time()

        self.test_project = ''
        self.test_job = ''
        self.case_excel_path = ''
        self.log_folder = ''
        self.dbTestJob = None
        self.Junit_xml = None

    def scene(self, scene_name, nexts, waiting_time, xpath, iss='n'):
        def logging_decorator(func):
            @wraps(func)
            def wrapped_function(previous_nodes_find, previous_scene, scene_yourself):
                func(previous_nodes_find, previous_scene, scene_yourself)

            mn = __import__(func.__module__).scene_module_name
            if not mn:
                # self.runner_logger(u'模块{}未定义scene_module_name'.format(func.__module__))
                mn = func.__module__

            if mn not in self.SceneModules:
                self.SceneModules[mn] = {}

            if scene_name in self.SceneModules[mn]:
                raise SceneRunnerRError(u'scene_name: "{}" 在模块: "{}" 中重复'.format(scene_name, mn))

            self.SceneModules[mn][scene_name] = (wrapped_function, nexts, waiting_time, xpath, iss)

            if iss == 's' or iss == 'e':
                self.SceneModules[mn]['_' + iss] = scene_name

            return wrapped_function

        return logging_decorator

    def scene_start(self, scene_name, nexts, waiting_time, xpath, iss='s'):
        return self.scene(scene_name, nexts, waiting_time, xpath, iss)

    def scene_end(self, scene_name, nexts, waiting_time, xpath, iss='e'):
        return self.scene(scene_name, nexts, waiting_time, xpath, iss)

    def runner_logger(self, msg, log_level=logging.INFO):

        if self.scene_now:
            msg = '[' + self.scene_now + '] - ' + msg
        if self.scene_module:
            msg = '[' + self.scene_module + '] - ' + msg
        logger.log(log_level, msg)

    def find_node_by_xpath(self, x, xx='', time_out=None):
        if None is not time_out:
            scene_app.flash_dump()
        else:
            time_out = 0

        for i in range(0, time_out + 1):
            r = scene_app.dump.xpath(x+xx)
            if r is None or len(r) == 0:
                self.flash_dump()
                time.sleep(1)
            else:
                return (r, i)
        return (None, i)

    def get_node_parents(self, n):
        if n.getparent() is None or n.getparent().tag == 'AbstractRoot':
            return ''
        # print n.getparent(),'n.getparent()'
        # print n.getparent().tag,'tag'
        return self.get_node_parents(n.getparent()) + '/' + n.getparent().tag

    def get_node_full_path(self, n):
        # print self.get_node_parents(n),'self.get_node_parents(n)'
        # print n.tag,'tag'
        return self.get_node_parents(n) + '/' + n.tag

    def node2gameobj(self, n):
        go = self.engine.find_elements_path(z2y(self.get_node_full_path(n)))
        if not go:
            raise SceneRunnerRError(u'无法通过以下路径找到GameObject:{}'.format(z2y(self.get_node_full_path(n))))

        # if len(go) > 1:
        #    raise SceneRunnerRWarning(u'通过以下路径找到的GameObject对象大于一个:{}'.format(n))
        return go[0]

    def press_node(self, n, press_time):
        return self.engine.press(self.node2gameobj(n), press_time)

    def find_gameobj_by_xpath(self, x,xx='', time_out=None):
        e, t = self.find_node_by_xpath(y2z(x),xx, time_out)
        if not e:
            raise SceneRunnerRError(u'无法通过xpath找到对象{} during {}'.format(x, time_out))
        if len(e) > 1:
            raise SceneRunnerRWarning(u'通过以下路径找到的xpath对象大于一个:{}'.format(x))

        return self.node2gameobj(e[0])

    def find_and_press(self, x, xx='', time_out=None, press_time=100):
        return self.engine.press(self.find_gameobj_by_xpath(x, xx, time_out), press_time)

    def find_and_swipe(self, x, x1,xx='', time_out=None, steps=20, duration=1000):
        return self.engine.swipe(
            self.find_gameobj_by_xpath(x,xx, time_out), self.find_gameobj_by_xpath(x1,xx, time_out), steps, duration)

    def find_and_press_random(self, x, xx='', time_out=None, press_time=100):
        s, t = self.find_node_by_xpath(y2z(x),xx, time_out)
        if not s:
            raise SceneRunnerRError(u'无法通过xpath找到对象{} during {} '.format(x, time_out))

        e = random.choice(s)

        self.engine.press(self.node2gameobj(e), press_time)

        return u'find_and_press_random在{}中选择了{}'.format(x, e.tag)

    def find_and_swipe_random(self, x, x1, time_out=None, steps=20, duration=1000):
        s, tt = self.find_node_by_xpath(y2z(x), time_out)
        if not s:
            raise SceneRunnerRError(u'无法通过xpath找到对象{} during {} '.format(x, time_out))

        e = random.choice(s)

        self.engine.swipe(self.node2gameobj(e), self.find_gameobj_by_xpath(x1, time_out), steps, duration)

        return u'find_and_press_random在{}中选择了{}'.format(x, e.tag)

    def screencap(self, p=''):

        p = self.log_folder + u'/{}-{}.png'.format(p, time.time())
        self.runner_logger(u'截图 {} '.format(p))

        mshell = ['exec-out', 'shell'][os.name != 'nt']

        with open(p, 'wb') as ps:
            if self.device:
                subprocess.call('adb -s ' + self.device + ' {} screencap -p'.format(mshell), stdout=ps, shell=True)
            else:
                subprocess.call('adb {} screencap -p'.format(mshell), stdout=ps, shell=True)

        return p

    def flash_dump(self):
        self.dump = treetraverse(etree.fromstring(get_wt_dump(self.engine)))[1]

    def runner_assert(self, identity, ass, iss=True, cf=None):
        if not self.scene_module or not self.scene_now:
            raise SceneRunnerRError(u'禁止在测试场景外使用断言')

        if identity not in self.Record:
            raise SceneRunnerRError(u'未定义用例标识符{}'.format(identity))

        if not cf:
            cf = inspect.currentframe()

        (frame, filename, line_number,
         function_name, lines, index) = inspect.getouterframes(cf)[1]

        filename = os.path.split(filename)[1] \
            .decode('utf-8')
        lines = unicode(lines[0].decode('utf-8'))

        if iss:
            p = self.screencap(identity)
            p = p.split('/')[-2] + '/' + p.split('/')[-1]
        else:
            p = ''

        du = time.time() - self.timer
        self.timer = time.time()

        self.Record[identity].append(
            [self.scene_module, self.scene_now, self.last_scene, bool(ass),
             datetime.fromtimestamp(self.timer - du), du, p, filename,
             line_number, function_name, lines])

        tx = u'断言 {} {}' \
             u'\n\t用例信息:' \
             u'\n\t\t模块：\n\t\t\t{}' \
             u'\n\t\t用例：\n\t\t\t{}' \
             u'\n\t\t步骤：\n\t\t\t{}' \
             u'\n\t\t预期结果：\n\t\t\t{}' \
             u'\n\t场景信息:' \
             u'\n\t\t{}:{}<-{}' \
             u'\n\t代码:' \
             u'\n\t\t@{}({}:{})' \
             u'\n\t\t\t=>{}'. \
            format(
            identity, [u'失败', u'成功'][bool(ass)],
            ur(self.Case[identity][0], u'\n', u'\n\t\t\t'),
            ur(self.Case[identity][1], u'\n', u'\n\t\t\t'),
            ur(self.Case[identity][2], u'\n', u'\n\t\t\t'),
            ur(self.Case[identity][3], u'\n', u'\n\t\t\t'),
            self.scene_module, self.scene_now, self.last_scene,
            function_name, filename, line_number, lines)

        self.Record[identity][-1].append(tx)

        # log
        for l in tx.split(u'\n'):
            self.runner_logger(l)

    def output_junit(self, path):
        self.Junit_xml = etree.Element("testsuites")

        for m in self.CaseTree:
            for c in self.CaseTree[m]:
                for i in self.CaseTree[m][c]:
                    e = etree.Element("testsuite", name='.'.join([m, c, i]))
                    for ri in range(0, len(self.Record[i])):
                        r = etree.Element('testcase', name=u'_{}'.format(ri), time=str(round(self.Record[i][ri][5], 2)))

                        if not self.Record[i][ri][3]:
                            er = etree.Element("error")
                            er.text = self.Record[i][ri][-1]
                            r.append(er)

                        e.append(r)

                    self.Junit_xml.append(e)

        etree.ElementTree(self.Junit_xml).write(
            open(path, 'w'), pretty_print=True, encoding='UTF-8', xml_declaration=True)

    def upload2db(self, test_project, test_job, record_info, db_uri, test_record_name, start, end):
        from db import TestProject, TestJob, TestRunRecord, TestCaseModule, TestCase, TestStepsSet, TestRealResult
        from db import connect_to_db
        from sqlobject import AND

        self.test_project = test_project
        self.test_job = test_job

        # 数据库准备操作
        try:
            connect_to_db(db_uri)
            self.dbTestJob = TestJob.select(AND(TestJob.j.TestProject, TestProject.q.Name == self.test_project,
                                                TestJob.q.Name == self.test_job)).getOne()
        except:
            raise
            # raise SceneRunnerRError(u'数据库初始化错误：{} {} {}'.format(db_uri, self.test_project, self.test_job))

        # rn = TestRunRecord.select(TestRunRecord.q.TestJob == self.dbTestJob)
        # n = '#{}'.format(int((rn.max(TestRunRecord.q.Name))[1:]) + 1) if rn.count() else '#0'

        rc = TestRunRecord(Name=test_record_name, Info=record_info, TestJob=self.dbTestJob, Start=start, End=end)
        for m in self.CaseTree:
            mm = TestCaseModule(Name=m, TestRunRecord=rc)
            for c in self.CaseTree[m]:
                cc = TestCase(Name=c, TestCaseModule=mm)
                for i in self.CaseTree[m][c]:
                    self.Case[i].append(TestStepsSet(Name=i, TestCase=cc, Steps=self.CaseTree[m][c][i][0],
                                                     ExpectedResult=self.CaseTree[m][c][i][1]))

        ks = ('SceneModule', 'Scene', 'LastScene', 'Assert', 'Start', 'During', 'TestScreenshotsID',
              'FileName', 'LineNumber', 'FunctionName', 'Line',)

        for identity in self.Record:
            for r in self.Record[identity]:
                TestRealResult(**dict(dict(zip(ks, tuple(r))), **{'TestStepsSet': self.Case[identity][-1]}))


                # return n

    def prepare(self, device, port,case_excel_path, log_folder, log_level=logging.INFO):
        # 初始化
        if log_level == logging.DEBUG:
            logger.setLevel(logging.DEBUG)
        else:
            self.check_scenes()

        self.device = device
        self.engine = get_engine(port)

        if not self.engine:
            raise SceneRunnerRError(u'引擎初始化错误，请检查ADB相关日志')

        self.case_excel_path = case_excel_path
        self.log_folder = log_folder

        # 用例导入
        try:
            # ws = xw.Book(self.case_excel_path).sheets[0]
            ws = load_workbook(filename=self.case_excel_path).active
        except:
            # raise SceneRunnerRError(u'用例Excel打开错误（注意不要使用中文文件名）：{} '.format(self.case_excel_path))
            raise

        # From F3
        r = 3
        while ws.cell(row=r, column=6).value:
            module_name, case_name, steps, expected_result, identity = [ws.cell(row=r, column=c).value for c in
                                                                        range(2, 7)]

            # for module_name, case_name, steps, expected_result, identity \
            #        in xw.Range(ws.range('B3'), ws.range('B3').end('down').end('right')).expand().value:

            if identity in self.Case:
                raise SceneRunnerRError(u'identity: "{}" 在Excel中重复'.format(identity))

            self.Case[identity] = [module_name, case_name, steps, expected_result]
            self.Record[identity] = []

            if module_name not in self.CaseTree:
                self.CaseTree[module_name] = {}
            if case_name not in self.CaseTree[module_name]:
                self.CaseTree[module_name][case_name] = {}
            if identity not in self.CaseTree[module_name][case_name]:
                self.CaseTree[module_name][case_name][identity] = (steps, expected_result)

            r += 1

        # ws.book.app.quit()

        return self

    def run_all(self):
        self.run_scenes(sorted(self.SceneModules.keys()))

    def run(self, case_excel_path, log_folder=u'./', log_level=logging.INFO, device=None,port=None):
        self.prepare(device, port,case_excel_path, log_folder, log_level)
        self.run_all()

    def run_scenes(self, scene_modules, wati_module=7):
        self.runner_logger(u'开始执行用例 共{}个'.format(len(self.SceneModules)))

        for self.scene_module in scene_modules:
            self.scenes = self.SceneModules[self.scene_module]
            self.runner_logger(u'开始执行用例  入口场景：{}  出口场景：{}'.format(self.scenes['_s'], self.scenes['_e']))

            nexts, wt, self.scene_now = ([self.scenes['_s']], wati_module, '_')

            self.timer = time.time()

            try:
                while 1:
                    if self.dump is None:
                        self.flash_dump()

                    self.runner_logger(u'寻找场景：{}'.format(repr(nexts).decode('unicode-escape')))
                    while 1:
                        ns = [self.dump.xpath(y2z(self.scenes[s][3])) for s in nexts]

                        rd = [(nexts[i], ns[i]) for i in range(0, len(nexts)) if ns[i]]

                        self.runner_logger(u'寻找匹配下游场景：{}  -  匹配：{}'.format(
                            repr(nexts).decode('unicode-escape'), repr([i[0] for i in rd]).decode('unicode-escape')),
                            logging.DEBUG)

                        if not rd:
                            if not wt:
                                # TODO NEED 更详细
                                if self.scene_now == '_':
                                    return self.scene_module
                                raise SceneRunnerRError(u'未找到匹配 倒计时超时')
                            else:
                                self.runner_logger(u'未找到匹配 倒计时：{}'.format(str(wt)), logging.DEBUG)
                                wt -= 1
                                time.sleep(1)
                                self.flash_dump()
                                continue
                        else:
                            break

                    if len(rd) > 1:
                        self.runner_logger(u'同时匹配一个以上场景,将优先选择', logging.WARNING)

                    self.runner_logger(u'匹配：{}'.format(repr([i[0] for i in rd]).decode('unicode-escape')))

                    self.last_scene = self.scene_now
                    self.scene_now, n = rd[0]

                    self.runner_logger(u'进入场景执行函数：{}'.format(self.scenes[self.scene_now][0].__name__))

                    self.scenes[self.scene_now][0](n, self.last_scene, self.scene_now)

                    self.runner_logger(u'退出场景执行函数：{}'.format(self.scenes[self.scene_now][0].__name__))
                    self.flash_dump()

                    if self.scene_now in [self.scenes['_e']]:
                        self.runner_logger(u'匹配出口场景 退出用例')
                        self.flash_dump()
                        break

                    nexts, wt = (self.scenes[self.scene_now][1], self.scenes[self.scene_now][2])
            except SceneRunnerRError as e:
                self.runner_logger(unicode(e), logging.ERROR)
                self.screencap('error')
                raise e
            except SceneRunnerRWarning as w:
                self.runner_logger(unicode(w), logging.WARNING)
                pass
            except SceneRunnerAssertionError as a:
                self.runner_logger(unicode(a), logging.WARNING)
                pass
            except:
                raise
            self.scene_module, self.scene_now = ('', '')
            self.runner_logger(u'结束执行用例 共{}个'.format(len(self.SceneModules)))

            return 0

    def check_scenes(self):

        logger.info(u'检查用例严谨性')
        try:
            for m in sorted(self.SceneModules):
                md = self.SceneModules[m]
                if '_s' not in md:
                    raise SceneRunnerRError(u'在模块:{}中未定义开始场景'.format(m))
                if '_e' not in md:
                    raise SceneRunnerRError(u'在模块:{}中未定义结束场景'.format(m))

                    # TODO NEED ADD MORE
        except SceneRunnerRError as e:
            logger.error(u'检查用例严谨性 - 失败')
            logger.error(e)
            raise e
        else:
            logger.info(u'检查用例严谨性 - 通过')

    def report(self):
        pass


scene_app = SceneRunner()

if __name__ == '__main__':
    scene_app.m = u'沙发上发呆'
    scene_app.screencap()
    # tf = '../temp/{0}.xml'.format(time.time())
    # save_wt_dump(open(tf, 'w'))
